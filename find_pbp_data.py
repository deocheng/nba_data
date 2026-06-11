#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
查找 Play-by-Play 数据
"""
import openpyxl

file_path = 'CSV/2026 final NYK vs SAS.xlsx'

print('=' * 80)
print('查找 Play-by-Play 数据')
print('=' * 80)

wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    print(f'\n工作表: {sheet_name}')
    print('=' * 80)
    
    ws = wb[sheet_name]
    
    # 查找所有包含时间的行
    print('搜索所有行，查找时间格式数据...')
    
    pbp_rows = []
    for row_num in range(1, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, min_col=1, max_col=30, values_only=True))[0]
        
        # 检查是否有时间格式（12:00, 11:30, etc）
        for cell in row:
            if cell and isinstance(cell, str):
                # 检查是否是时间格式
                if ':' in str(cell) and any(f'{i}:' in str(cell) for i in range(12)):
                    pbp_rows.append((row_num, row))
                    break
    
    print(f'找到 {len(pbp_rows)} 行包含时间格式数据')
    
    if pbp_rows:
        print('\n前 30 行 Play-by-Play 数据:')
        for row_num, row in pbp_rows[:30]:
            row_preview = [str(cell)[:20] if cell else '' for cell in row[:15]]
            print(f'Row {row_num}: {repr(row_preview)}')

print('\n搜索更宽的列范围...')
