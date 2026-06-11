#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析 Excel 文件的详细内容
"""
import openpyxl
from openpyxl import load_workbook

file_path = 'CSV/2026 final NYK vs SAS.xlsx'

print('=' * 80)
print('详细分析: 2026 final NYK vs SAS.xlsx')
print('=' * 80)
print()

wb = load_workbook(file_path)

for sheet_name in wb.sheetnames:
    print(f'工作表: {sheet_name}')
    print('=' * 80)
    
    ws = wb[sheet_name]
    
    # 打印前100行，前30列的内容
    print(f'前100行，前30列:')
    print()
    
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=30, values_only=True), 1):
        # 过滤全空的行
        if all(cell is None for cell in row):
            continue
            
        # 打印行号和内容
        row_str = [str(cell) if cell is not None else '' for cell in row]
        print(f'Row {row_num}: {repr(row_str)}')
        
        # 找到 play-by-play 的开始（通常包含 Quarter 或 1st 等信息）
        row_lower = str(row).lower()
        if any(keyword in row_lower for keyword in ['play by play', 'play-by-play', 'quarter', 'q1', '1st', 'time', 'visitor', 'home']):
            print(f'  >>> 可能是 Play-by-Play 数据开始的地方！')
            print()
    
    print()
