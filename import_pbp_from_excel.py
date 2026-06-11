#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Excel 导入 Play-by-Play 数据
"""
import sys
import os
import openpyxl
import re
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from data_importer.database import DatabaseManager
from data_importer.config import DatabaseConfig

def parse_pbp_row(row):
    """解析 Play-by-Play 单行"""
    
    # 处理第一个单元格（可能是 datetime.time 或字符串）
    first_cell = row[0]
    
    # 如果是 datetime.time 对象（有 hour 和 minute 属性）
    if first_cell and hasattr(first_cell, 'hour') and hasattr(first_cell, 'minute'):
        time_str = f"{first_cell.hour:02d}:{first_cell.minute:02d}:{first_cell.second:02d}"
        game_clock = f"{first_cell.minute}:{first_cell.second:02d}"
    elif first_cell and isinstance(first_cell, str):
        time_str = first_cell
        match = re.match(r'(\d{2}):(\d{2}):(\d{2})', time_str)
        if not match:
            return None
        game_clock = f"{match.group(1)}:{match.group(2)}"
    else:
        return None
    
    cells = [str(cell) if cell else '' for cell in row]
    
    # 提取客队和主队得分
    visitor_score = None
    home_score = None
    
    # 检查所有单元格中的比分
    score_pattern = re.compile(r'(\d+)-(\d+)')
    for cell in cells:
        if score_pattern.search(cell):
            scores = score_pattern.findall(cell)
            if scores:
                # 取最后一个比分（最新的）
                last_score = scores[-1]
                visitor_score = int(last_score[0])
                home_score = int(last_score[1])
                break
    
    # 提取动作描述
    actions = []
    for cell in cells[1:]:
        if cell and cell.strip():
            # 清理动作文本
            action = cell.strip()
            if len(action) > 5:  # 过滤掉太短的文本
                actions.append(action)
    
    return {
        'time': time_str,
        'game_clock': game_clock,
        'visitor_score': visitor_score,
        'home_score': home_score,
        'actions': actions
    }

def import_pbp_from_sheet(db_manager, ws, sheet_name, game_id):
    """从工作表导入 Play-by-Play 数据"""
    
    print(f'\n导入 Play-by-Play 数据: {sheet_name}')
    print('=' * 60)
    
    # 找到 Play-by-Play 开始的行
    pbp_start = None
    for row_num in range(1, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, min_col=1, max_col=10, values_only=True))[0]
        
        first_cell = row[0]
        
        # 检查是否是时间格式（字符串或datetime.time对象）
        if first_cell:
            # 如果是 datetime.time 对象
            if hasattr(first_cell, 'hour') and hasattr(first_cell, 'minute'):
                pbp_start = row_num
                print(f'找到 Play-by-Play 数据开始: Row {row_num} (datetime.time)')
                break
            # 如果是字符串
            elif isinstance(first_cell, str) and re.match(r'\d{2}:\d{2}:\d{2}', str(first_cell)):
                pbp_start = row_num
                print(f'找到 Play-by-Play 数据开始: Row {row_num} (str)')
                break
    
    if not pbp_start:
        print('⚠️  未找到 Play-by-Play 数据')
        return 0
    
    print(f'Play-by-Play 数据从 Row {pbp_start} 开始')
    
    # 获取游戏 ID
    game_sql = "SELECT game_id FROM games WHERE game_date >= '2026-06-01' ORDER BY game_id DESC LIMIT 1"
    result = db_manager.fetch_one(game_sql)
    if not result:
        print('⚠️  未找到对应的比赛记录')
        return 0
    
    db_game_id = result[0]
    print(f'游戏 ID: {db_game_id}')
    
    # 解析所有 Play-by-Play 行
    pbp_records = []
    for row_num in range(pbp_start, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, min_col=1, max_col=20, values_only=True))[0]
        
        parsed = parse_pbp_row(row)
        if parsed:
            # 分离客队和主队动作
            visitor_action = None
            home_action = None
            
            # 根据比分变化判断是客队还是主队得分
            actions_text = ' | '.join(parsed['actions'][:3])
            if parsed['visitor_score'] and parsed['home_score']:
                # 简单判断：如果有"makes"、"misses"等关键词
                for action in parsed['actions']:
                    if 'NYK' in action or 'New York' in action:
                        visitor_action = action
                    elif 'SAS' in action or 'San Antonio' in action:
                        home_action = action
            
            pbp_records.append({
                'game_id': db_game_id,
                'season_end': 2026,
                'period': 1,  # 默认第1节，后续可以改进检测
                'game_clock': parsed['game_clock'],
                'visitor_action': visitor_action,
                'home_action': home_action,
                'visitor_score': parsed['visitor_score'],
                'home_score': parsed['home_score'],
                'score_change': None,
                'row_num': row_num
            })
    
    print(f'解析到 {len(pbp_records)} 条 Play-by-Play 记录')
    
    if pbp_records:
        # 保存到数据库
        # 注意：play_by_play 表可能有不同的结构，需要检查
        try:
            inserted = db_manager.insert_data('play_by_play', pbp_records, batch_size=100)
            print(f'✅ 成功导入 {inserted} 条 Play-by-Play 记录')
            return inserted
        except Exception as e:
            print(f'⚠️  导入失败: {e}')
            
            # 尝试创建简化版表
            try:
                create_sql = """
                CREATE TABLE IF NOT EXISTS excel_pbp_data (
                    id SERIAL PRIMARY KEY,
                    game_id INTEGER,
                    time VARCHAR(20),
                    game_clock VARCHAR(20),
                    visitor_score INTEGER,
                    home_score INTEGER,
                    actions TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
                """
                db_manager.execute(create_sql)
                inserted = db_manager.insert_data('excel_pbp_data', pbp_records, batch_size=100)
                print(f'✅ 保存到 excel_pbp_data 表: {inserted} 条')
                return inserted
            except Exception as e2:
                print(f'⚠️  保存失败: {e2}')
                return 0
    
    return 0

def main():
    print('=' * 80)
    print('导入 Play-by-Play 数据')
    print('=' * 80)
    
    db_config = DatabaseConfig()
    db_manager = DatabaseManager(db_config)
    
    excel_file = 'CSV/2026 final NYK vs SAS.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    
    total_imported = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = import_pbp_from_sheet(db_manager, ws, sheet_name, sheet_name)
        total_imported += count
    
    print('\n' + '=' * 80)
    print(f'✅ 总共导入 {total_imported} 条 Play-by-Play 记录')
    print('=' * 80)
    
    db_manager.close()

if __name__ == '__main__':
    main()
