#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Excel 文件导入数据到数据库
"""
import sys
import os
import json
import openpyxl
from datetime import datetime
import re

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from data_importer.database import DatabaseManager
from data_importer.config import DatabaseConfig

def main():
    print('=' * 80)
    print('导入 Excel 数据到数据库')
    print('=' * 80)
    print()
    
    # 初始化数据库
    db_config = DatabaseConfig()
    db_manager = DatabaseManager(db_config)
    
    # 检查并创建需要的表/列
    create_tables_if_needed(db_manager)
    
    # Excel 文件路径
    excel_file = 'CSV/2026 final NYK vs SAS.xlsx'
    
    if not os.path.exists(excel_file):
        print(f'文件不存在: {excel_file}')
        db_manager.close()
        return
    
    wb = openpyxl.load_workbook(excel_file)
    
    for sheet_name in wb.sheetnames:
        print(f'\n处理工作表: {sheet_name}')
        print('=' * 60)
        
        ws = wb[sheet_name]
        
        # 解析比赛信息
        game_info = parse_game_info_from_sheet(ws, sheet_name)
        
        if game_info:
            print(f'比赛信息: {game_info["visitor_team"]} vs {game_info["home_team"]}')
            print(f'比分: {game_info["visitor_score"]} - {game_info["home_score"]}')
            print(f'日期: {game_info["date"]}')
            
            # 保存到 game_metadata 表
            save_game_to_database(db_manager, game_info)
            
            # 解析球员统计
            player_stats = parse_player_stats_from_sheet(ws)
            
            if player_stats:
                print(f'找到 {len(player_stats)} 个球员统计')
                
                # 保存球员统计到数据库
                save_player_stats_to_database(db_manager, game_info, player_stats)
                
                print(f'✅ 球员统计已保存')
        else:
            print('⚠️  未能解析比赛信息')
    
    print()
    print('=' * 80)
    print('✅ 导入完成')
    print('=' * 80)
    
    db_manager.close()

def create_tables_if_needed(db_manager):
    """添加缺失的列"""
    
    # 检查 game_metadata 表是否需要 location 和 series_summary 列
    check_location_col_sql = """
    SELECT 1 FROM information_schema.columns 
    WHERE table_name = 'game_metadata' AND column_name = 'location'
    """
    location_exists = db_manager.fetch_one(check_location_col_sql)
    
    if not location_exists:
        add_location_sql = "ALTER TABLE game_metadata ADD COLUMN location VARCHAR(200)"
        add_series_summary_sql = "ALTER TABLE game_metadata ADD COLUMN series_summary TEXT"
        db_manager.execute(add_location_sql)
        db_manager.execute(add_series_summary_sql)
        print('✅ 为 game_metadata 添加了 location 和 series_summary 列')
    
    print('✅ 表结构检查/调整完成')

def parse_game_info_from_sheet(ws, sheet_name):
    """从工作表解析比赛信息"""
    
    # 从第一行获取标题
    game_info = {
        'game_id': '',
        'season_end': 2026,
        'visitor_team': '',
        'home_team': '',
        'visitor_score': None,
        'home_score': None,
        'date': None,
        'location': '',
        'series_summary': '',
        'sheet_name': sheet_name
    }
    
    # 搜索前 50 行获取比赛信息
    for row_num in range(1, 51):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, min_col=1, max_col=10, values_only=True))[0]
        
        if row[0]:
            cell_val = str(row[0])
            
            # 提取日期和比赛标识
            if 'Game' in cell_val and 'Jun' in cell_val:
                game_info['date'] = '2026-06-' + cell_val.split(', ')[1].split(' ')[1]
                
            # 提取球队和比分
            elif cell_val in ['NYK', 'SAS']:
                if not game_info['visitor_team']:
                    game_info['visitor_team'] = 'New York Knicks' if cell_val == 'NYK' else 'San Antonio Spurs'
                    game_info['visitor_score'] = int(row[1]) if row[1] else None
                else:
                    game_info['home_team'] = 'New York Knicks' if cell_val == 'NYK' else 'San Antonio Spurs'
                    game_info['home_score'] = int(row[1]) if row[1] else None
                    
            # 提取比赛地点
            elif 'Frost Bank Center' in cell_val or 'San Antonio' in cell_val:
                game_info['location'] = cell_val
                
            # 提取 series summary
            elif 'Series Summary' in cell_val:
                game_info['series_summary'] = f'G{sheet_name[1:]}: {game_info["visitor_score"]}-{game_info["home_score"]}'
    
    # 生成 game_id
    if 'G1' in sheet_name:
        game_info['game_id'] = '202606030SAS'
    elif 'G2' in sheet_name:
        game_info['game_id'] = '202606050SAS'
        
    return game_info

def parse_player_stats_from_sheet(ws):
    """从工作表解析球员统计"""
    
    player_stats = []
    in_knicks = False
    in_spurs = False
    in_reserves = False
    
    # 无效球员名字列表
    invalid_names = [
        "Starters",
        "Reserves",
        "Team Totals",
        "Share &",
        "Modify,",
        "Get as",
        "Get table",
        "Get Link",
        "About",
        "Video:",
        "Data",
        "Glossary",
        "Advanced",
    ]
    
    # 搜索前 200 行
    for row_num in range(1, 201):
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, min_col=1, max_col=25, values_only=True))[0]
        
        if not any(row):
            continue
            
        first_cell = str(row[0]) if row[0] else ''
        
        # 检查球队标志
        if 'New York Knicks' in first_cell and 'Stats' in first_cell:
            in_knicks = True
            in_spurs = False
            continue
        elif 'San Antonio Spurs' in first_cell and 'Stats' in first_cell:
            in_spurs = True
            in_knicks = False
            continue
        elif 'Reserves' == first_cell:
            in_reserves = True
            continue
        elif 'Team Totals' == first_cell:
            in_knicks = False
            in_spurs = False
            in_reserves = False
            continue
        elif 'Starters' == first_cell:
            in_reserves = False
            continue
        
        # 检查是否是球员行
        if (in_knicks or in_spurs) and row[0]:
            # 判断是否是无效球员
            if any(invalid in first_cell for invalid in invalid_names):
                continue
            
            # 判断是否是 valid 球员：包含至少一个空格，或者是常见的名字模式
            name_parts = first_cell.split()
            if len(name_parts) < 2:
                continue
                
            # 检查是否是"Did Not Play"
            if len(row) > 1 and row[1]:
                if 'Did Not Play' in str(row[1]):
                    continue
            
            # 解析统计数据
            try:
                # 解析分钟
                mp_str = str(row[1]) if row[1] else None
                minutes = parse_minutes(mp_str)
                
                stat = {
                    'team': 'New York Knicks' if in_knicks else 'San Antonio Spurs',
                    'player_name': first_cell,
                    'is_starter': not in_reserves,
                    'mp': mp_str,
                    'minutes': minutes,
                    'fg': int(row[2]) if row[2] is not None and str(row[2]).strip() != '' else None,
                    'fga': int(row[3]) if row[3] is not None and str(row[3]).strip() != '' else None,
                    'fg_pct': float(row[4]) if row[4] is not None and str(row[4]).strip() != '' else None,
                    'three_p': int(row[5]) if row[5] is not None and str(row[5]).strip() != '' else None,
                    'three_pa': int(row[6]) if row[6] is not None and str(row[6]).strip() != '' else None,
                    'three_p_pct': float(row[7]) if row[7] is not None and str(row[7]).strip() != '' else None,
                    'ft': int(row[8]) if row[8] is not None and str(row[8]).strip() != '' else None,
                    'fta': int(row[9]) if row[9] is not None and str(row[9]).strip() != '' else None,
                    'ft_pct': float(row[10]) if row[10] is not None and str(row[10]).strip() != '' else None,
                    'orb': int(row[11]) if row[11] is not None and str(row[11]).strip() != '' else None,
                    'drb': int(row[12]) if row[12] is not None and str(row[12]).strip() != '' else None,
                    'trb': int(row[13]) if row[13] is not None and str(row[13]).strip() != '' else None,
                    'ast': int(row[14]) if row[14] is not None and str(row[14]).strip() != '' else None,
                    'stl': int(row[15]) if row[15] is not None and str(row[15]).strip() != '' else None,
                    'blk': int(row[16]) if row[16] is not None and str(row[16]).strip() != '' else None,
                    'tov': int(row[17]) if row[17] is not None and str(row[17]).strip() != '' else None,
                    'pf': int(row[18]) if row[18] is not None and str(row[18]).strip() != '' else None,
                    'pts': int(row[19]) if row[19] is not None and str(row[19]).strip() != '' else None,
                    'gmsc': float(row[20]) if len(row) > 20 and row[20] is not None and str(row[20]).strip() != '' else None,
                    'plus_minus': int(row[21]) if len(row) > 21 and row[21] is not None and str(row[21]).strip() != '' else None,
                }
                
                player_stats.append(stat)
                
            except Exception as e:
                pass
    
    return player_stats

def parse_minutes(mp_str):
    """解析时间字符串为分钟数"""
    if not mp_str or 'Did Not' in mp_str:
        return None
        
    try:
        # 格式: 38:42, 30, 12, etc
        mp_str = str(mp_str).strip()
        
        if ':' in mp_str:
            parts = mp_str.split(':')
            minutes = int(parts[0])
            seconds = int(parts[1])
            return minutes + seconds / 60
        else:
            # 纯数字
            return float(mp_str)
    except:
        return None

def get_team_id(db_manager, team_name):
    """获取球队 ID"""
    sql = "SELECT team_id FROM teams WHERE name = %s OR abbreviation = %s"
    team = db_manager.fetch_one(sql, (team_name, team_name))
    if team:
        return team[0]
    return None

def get_or_create_player(db_manager, player_name):
    """获取或创建球员"""
    sql = "SELECT player_id FROM players WHERE name = %s"
    player = db_manager.fetch_one(sql, (player_name,))
    
    if player:
        return player[0]
    
    # 创建球员
    insert_sql = "INSERT INTO players (name) VALUES (%s) RETURNING player_id"
    result = db_manager.fetch_one(insert_sql, (player_name,))
    return result[0] if result else None

def save_game_to_database(db_manager, game_info):
    """保存比赛信息到数据库"""
    
    # 检查是否已存在
    check_sql = "SELECT id FROM game_metadata WHERE game_id = %s"
    existing = db_manager.fetch_one(check_sql, (game_info['game_id'],))
    
    if existing:
        # 更新
        update_sql = """
        UPDATE game_metadata 
        SET visitor_team = %s, home_team = %s,
            visitor_score = %s, home_score = %s,
            game_date = %s, location = %s, series_summary = %s
        WHERE game_id = %s
        """
        db_manager.execute(update_sql, (
            game_info['visitor_team'],
            game_info['home_team'],
            game_info['visitor_score'],
            game_info['home_score'],
            game_info['date'],
            game_info['location'],
            game_info['series_summary'],
            game_info['game_id']
        ))
        print(f'✅ 更新了比赛信息: {game_info["game_id"]}')
    else:
        # 插入
        insert_sql = """
        INSERT INTO game_metadata (
            game_id, season_end, visitor_team, home_team, 
            visitor_score, home_score, game_date, location, series_summary
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        db_manager.execute(insert_sql, (
            game_info['game_id'],
            game_info['season_end'],
            game_info['visitor_team'],
            game_info['home_team'],
            game_info['visitor_score'],
            game_info['home_score'],
            game_info['date'],
            game_info['location'],
            game_info['series_summary']
        ))
        print(f'✅ 新增了比赛信息: {game_info["game_id"]}')

def get_or_create_game(db_manager, game_info):
    """获取或创建比赛记录"""
    # 获取球队 ID
    visitor_team_id = get_team_id(db_manager, game_info['visitor_team'])
    home_team_id = get_team_id(db_manager, game_info['home_team'])
    
    if not visitor_team_id or not home_team_id:
        print(f'⚠️  无法找到球队 ID: visitor={game_info["visitor_team"]}, home={game_info["home_team"]}')
        return None
        
    # 检查游戏是否存在
    check_sql = "SELECT game_id FROM games WHERE game_date = %s AND home_team_id = %s"
    existing = db_manager.fetch_one(check_sql, (game_info['date'], home_team_id))
    
    if existing:
        return existing[0]
        
    # 创建游戏
    insert_sql = """
    INSERT INTO games (
        game_date, season, home_team_id, away_team_id,
        home_score, away_score, winner_team_id, status
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    RETURNING game_id
    """
    winner_id = home_team_id if (game_info['home_score'] > game_info['visitor_score']) else visitor_team_id
    
    result = db_manager.fetch_one(insert_sql, (
        game_info['date'],
        str(game_info['season_end']),
        home_team_id,
        visitor_team_id,
        game_info['home_score'],
        game_info['visitor_score'],
        winner_id,
        'Final'
    ))
    
    return result[0] if result else None
    
def save_player_stats_to_database(db_manager, game_info, player_stats):
    """保存球员统计到数据库"""
    
    # 获取比赛 ID
    game_id_for_ref = get_or_create_game(db_manager, game_info)
    
    if not game_id_for_ref:
        print(f'⚠️  无法获取/创建比赛记录')
        return
    
    # 准备批量数据
    batch_data = []
    
    for stat in player_stats:
        # 获取/创建球员
        player_id = get_or_create_player(db_manager, stat['player_name'])
        
        # 获取球队 ID
        team_id = get_team_id(db_manager, stat['team'])
        
        if not player_id or not team_id:
            print(f'⚠️  无法处理球员: {stat["player_name"]} (team: {stat["team"]})')
            continue
        
        # 检查是否已存在
        check_sql = "SELECT id FROM player_game_stats WHERE game_id = %s AND player_id = %s"
        existing = db_manager.fetch_one(check_sql, (game_id_for_ref, player_id))
        
        if existing:
            print(f'  ℹ️  球员已存在，跳过: {stat["player_name"]}')
            continue
            
        # 准备数据字典
        record = {
            'game_id': game_id_for_ref,
            'player_id': player_id,
            'team_id': team_id,
            'minutes': stat['minutes'],
            'fg': stat['fg'],
            'fga': stat['fga'],
            'fg_pct': stat['fg_pct'],
            'three_p': stat['three_p'],
            'three_pa': stat['three_pa'],
            'three_p_pct': stat['three_p_pct'],
            'ft': stat['ft'],
            'fta': stat['fta'],
            'ft_pct': stat['ft_pct'],
            'orb': stat['orb'],
            'drb': stat['drb'],
            'trb': stat['trb'],
            'ast': stat['ast'],
            'stl': stat['stl'],
            'blk': stat['blk'],
            'tov': stat['tov'],
            'pf': stat['pf'],
            'pts': stat['pts'],
            'gmsc': stat['gmsc'],
            'plus_minus': stat['plus_minus'],
        }
        
        batch_data.append(record)
        print(f'  ✅ 准备球员: {stat["player_name"]} ({stat["pts"]} 分)')
    
    if batch_data:
        # 批量插入
        inserted = db_manager.insert_data('player_game_stats', batch_data, batch_size=10)
        print(f'\n✅ 批量插入了 {inserted} 条球员记录')

if __name__ == '__main__':
    main()
