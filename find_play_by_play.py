#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
在 Excel 中查找 Play-by-Play 数据
"""
import openpyxl

file_path = 'CSV/2026 final NYK vs SAS.xlsx'

print('=' * 80)
print('查找 Play-by-Play 数据')
print('=' * 80)
print()

wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    print(f'工作表: {sheet_name}')
    print('=' * 80)
    
    ws = wb[sheet_name]
    
    # 搜索更多行，找到可能的 Play-by-Play 表格
    play_by_play_start = None
    
    print('搜索前200行，查找 Play-by-Play 表格:')
    print()
    
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=30, values_only=True), 1):
        row_str = str(row).lower()
        
        # 检查 Play-by-Play 的特征
        keywords = [
            'play by play',
            'play-by-play',
            '12:00',
            '11:00',
            '10:00',
            '09:00',
            '0:00',
            'timeout',
            'quarter',
            'visitor',
            'visitors',
            'home',
            'score'
        ]
        
        if any(k in row_str for k in keywords):
            print(f'Row {row_num}: {repr([str(cell)[:50] if cell is not None else "" for cell in row[:10]])}')
            
            # 如果找到多个相关行，可能就是 Play-by-Play 开始的地方
            if play_by_play_start is None:
                play_by_play_start = row_num
                print(f'  >>> 可能是 Play-by-Play 数据开始的地方（行 {row_num}）')
    
    print()
    
    # 如果找到了开始位置，查看周围的数据
    if play_by_play_start:
        print(f'查看从行 {play_by_play_start} 开始的 100 行:')
        print('=' * 80)
        for row_num in range(play_by_play_start, play_by_play_start + 100):
            row = list(ws.iter_rows(min_row=row_num, max_row=row_num, min_col=1, max_col=20, values_only=True))[0]
            row_str = [str(cell) if cell is not None else '' for cell in row]
            print(f'Row {row_num}: {repr(row_str)}')
        print()
